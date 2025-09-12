# === AUTO-GENERATED INSERT PATCH ===
# target_module: excel_eval.py
# insert_blocks follow; keep formatting unchanged

# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: func excel_round ---
def excel_round(value: float, digits: int = 0) -> float:
    """Excel-ROUND: half away from zero (anders als Python round)."""
    if digits < 0:
        factor = 10 ** (-digits)
        return math.copysign(math.floor(abs(value) / factor + 0.5) * factor, value)
    factor = 10 ** digits
    return math.copysign(math.floor(abs(value) * factor + 0.5) / factor, value)
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: func excel_serial_to_date ---
def excel_serial_to_date(serial: float, date1904: bool = False):
    """Excel-Seriendatum -> datetime.date (inkl. 1900-Bug/1904-System)."""
    from datetime import date, timedelta
    if date1904:
        base = date(1904, 1, 1)
        return base + timedelta(days=int(serial))
    base = date(1899, 12, 30)  # kompensiert Excel 1900-Fehler
    return base + timedelta(days=int(serial))
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: func _value_type_of ---
def _value_type_of(v: Any) -> str:
    if v is None:
        return "blank"
    if isinstance(v, bool):
        return "bool"
    if isinstance(v, (int, float)):
        return "number"
    if isinstance(v, str):
        # sehr einfache Fehlererkennung; xlwings gibt native Fehlerobjekte auf Windows,
        # headless gibt Strings wie "#N/A". Wir normalisieren auf String.
        if v.startswith("#"):
            return "error"
        return "text"
    # Excel kann z.B. datetime liefern:
    try:
        import datetime as _dt
        if isinstance(v, (_dt.date, _dt.datetime)):
            return "date"
    except Exception:
        pass
    return type(v).__name__
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: func scan_formulas_openpyxl ---
def scan_formulas_openpyxl(path: str | Path, limit: Optional[int] = None) -> List[Tuple[str, str, str]]:
    """
    Liest die Datei, sammelt alle Formelzellen (Sheet, Address, Formula).
    data_only=False => echte Formeln, nicht Cache-Werte.
    """
    from openpyxl import load_workbook
    path = str(path)
    wb = load_workbook(path, data_only=False, read_only=True)
    out: List[Tuple[str, str, str]] = []
    for ws in wb.worksheets:
        dim = ws.calculate_dimension()  # begrenzt iter_rows auf "Used Range"
        # iter_rows mit values_only=False liefert Cell-Objekte
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    out.append((ws.title, cell.coordinate, v))
                    if limit and len(out) >= limit:
                        wb.close()
                        return out
    wb.close()
    return out
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: func _engine_auto ---
def _engine_auto(path: str | Path, prefer: str = "xlwings"):
    """
    Liefert (engine_name, engine_instance). prefer in {"xlwings","headless"}.
    """
    path = Path(path)
    if prefer == "xlwings":
        try:
            eng = ExcelEngineXlwings(path)
            return "xlwings", eng
        except Exception:
            pass
        # fallback
        eng = ExcelEngineHeadless(path)
        return "headless", eng
    else:
        try:
            eng = ExcelEngineHeadless(path)
            return "headless", eng
        except Exception:
            # fallback
            eng = ExcelEngineXlwings(path)
            return "xlwings", eng
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: func _evaluate_many ---
def _evaluate_many(path: str | Path, refs: List[Tuple[str, str, str]],
                   engine: str = "auto") -> List[FormulaRow]:
    """
    refs: Liste (sheet, address, formula)
    engine: "auto" | "xlwings" | "headless"
    """
    rows: List[FormulaRow] = []
    if engine == "auto":
        engine_name, eng = _engine_auto(path, prefer="xlwings")
    elif engine == "xlwings":
        eng = ExcelEngineXlwings(path)
        engine_name = "xlwings"
    elif engine == "headless":
        eng = ExcelEngineHeadless(path)
        engine_name = "headless"
    else:
        raise ValueError("engine must be auto|xlwings|headless")

    try:
        for (sheet, address, formula) in refs:
            try:
                if engine_name == "xlwings":
                    v = eng.value(sheet, address)
                else:
                    v = eng.evaluate(f"{sheet}!{address}")
            except Exception as e:
                v = f"#ERR:{type(e).__name__}:{e}"
            rows.append(FormulaRow(
                sheet=sheet,
                address=address,
                formula=formula,
                value=v,
                value_type=_value_type_of(v),
            ))
    finally:
        if engine_name == "xlwings":
            eng.close()
    return rows
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: func export_formulas_to_csv ---
def export_formulas_to_csv(path: str | Path, out_csv: str | Path,
                           limit: Optional[int] = None, engine: str = "auto") -> Path:
    """
    Exportiert Formelzellen (mit berechnetem Wert) als CSV.
    Spalten: sheet,address,formula,value,value_type
    """
    refs = scan_formulas_openpyxl(path, limit=limit)
    rows = _evaluate_many(path, refs, engine=engine)
    out_csv = Path(out_csv)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "address", "formula", "value", "value_type"])
        for r in rows:
            w.writerow([r.sheet, r.address, r.formula, r.value, r.value_type])
    return out_csv
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: func export_formulas_to_xlsx ---
def export_formulas_to_xlsx(path: str | Path, out_xlsx: str | Path,
                            limit: Optional[int] = None, engine: str = "auto") -> Path:
    """
    Exportiert Formelzellen (mit berechnetem Wert) als .xlsx
    """
    import pandas as pd
    refs = scan_formulas_openpyxl(path, limit=limit)
    rows = _evaluate_many(path, refs, engine=engine)
    df = pd.DataFrame([{
        "sheet": r.sheet,
        "address": r.address,
        "formula": r.formula,
        "value": r.value,
        "value_type": r.value_type,
    } for r in rows])
    out_xlsx = Path(out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(out_xlsx, index=False)
    return out_xlsx
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: func batch_export ---
def batch_export(
    base: str | Path,
    pattern: str = "**/*.xlsx",
    out_dir: str | Path = "exports",
    limit: Optional[int] = None,
    engine: str = "auto",
    to: str = "csv",   # "csv" | "xlsx"
) -> List[Tuple[Path, Path]]:
    """
    Sucht Dateien via Glob und exportiert pro Datei die Formel-Ergebnisse.
    Rückgabe: Liste [(infile, outfile), ...]
    """
    base = Path(base)
    out_dir = Path(out_dir)
    results: List[Tuple[Path, Path]] = []
    for p in base.rglob("*"):
        if not p.is_file():
            continue
        if p.suffix.lower() != ".xlsx":
            continue
        if p.name.startswith("~$"):  # temporäre Excel-Dateien
            continue
        rel = p.relative_to(base)
        stem = p.stem
        out_subdir = out_dir / rel.parent
        out_subdir.mkdir(parents=True, exist_ok=True)
        if to == "csv":
            out_file = out_subdir / f"{stem}_formulas.csv"
            export_formulas_to_csv(p, out_file, limit=limit, engine=engine)
        else:
            out_file = out_subdir / f"{stem}_formulas.xlsx"
            export_formulas_to_xlsx(p, out_file, limit=limit, engine=engine)
        results.append((p, out_file))
    return results
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: func safety_check_parity ---
def safety_check_parity(
    path: str | Path,
    sample_limit: Optional[int] = 1000,
    abs_tol: float = 1e-9,
    rel_tol: float = 1e-9,
    out_report_csv: Optional[str | Path] = None,
) -> Dict[str, Any]:
    """
    Vergleicht xlwings (Excel) vs. headless (xlcalculator).
    Liefert Zusammenfassung + optional CSV-Report mit Abweichungen.
    """
    refs = scan_formulas_openpyxl(path, limit=sample_limit)

    # Excel
    excel_rows = _evaluate_many(path, refs, engine="xlwings")
    # Headless
    head_rows = _evaluate_many(path, refs, engine="headless")

    mismatches: List[Mismatch] = []

    def _num(x) -> Optional[float]:
        return float(x) if isinstance(x, (int, float)) else None

    for er, hr in zip(excel_rows, head_rows):
        ref = f"{er.sheet}!{er.address}"
        ev = er.value
        hv = hr.value

        note = ""
        abs_d = None
        rel_d = None

        # Fehler/Strings direkt vergleichen
        if isinstance(ev, str) or isinstance(hv, str):
            if ev != hv:
                mismatches.append(Mismatch(ref, ev, hv, None, None, "string_mismatch"))
            continue

        # Zahlenvergleich mit Toleranz
        if isinstance(ev, (int, float)) and isinstance(hv, (int, float)):
            evf = float(ev)
            hvf = float(hv)
            abs_d = abs(evf - hvf)
            rel_d = abs_d / (abs(evf) + 1e-30)
            if not (abs_d <= abs_tol or rel_d <= rel_tol):
                mismatches.append(Mismatch(ref, ev, hv, abs_d, rel_d, "float_delta"))
            continue

        # direkte Gleichheit für Resttypen
        if ev != hv:
            mismatches.append(Mismatch(ref, ev, hv, None, None, "value_mismatch"))

    summary = {
        "file": str(path),
        "checked_cells": len(refs),
        "mismatches": len(mismatches),
        "abs_tol": abs_tol,
        "rel_tol": rel_tol,
    }

    if out_report_csv:
        out_report_csv = Path(out_report_csv)
        out_report_csv.parent.mkdir(parents=True, exist_ok=True)
        with out_report_csv.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["ref", "excel_value", "headless_value", "abs_delta", "rel_delta", "note"])
            for m in mismatches:
                w.writerow([m.ref, m.excel_value, m.headless_value, m.abs_delta, m.rel_delta, m.note])
        summary["report"] = str(out_report_csv)

    return summary
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: func _print_json ---
def _print_json(obj):
    print(json.dumps(obj, ensure_ascii=False, indent=2, default=str))
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: func main_cli ---
def main_cli():
    """
    Kleine CLI:
    python -m excel_eval export --in file.xlsx --out out.csv --engine auto --limit 0
    python -m excel_eval batch --base ./input --out ./exports --to csv
    python -m excel_eval safety --in file.xlsx --abs 1e-9 --rel 1e-9 --report report.csv
    """
    import argparse
    p = argparse.ArgumentParser("excel_eval")
    sub = p.add_subparsers(dest="cmd", required=True)

    p_exp = sub.add_parser("export")
    p_exp.add_argument("--in", dest="infile", required=True)
    p_exp.add_argument("--out", dest="outfile", required=True)
    p_exp.add_argument("--engine", choices=["auto","xlwings","headless"], default="auto")
    p_exp.add_argument("--limit", type=int, default=None)
    p_exp.add_argument("--fmt", choices=["csv","xlsx"], default="csv")

    p_bat = sub.add_parser("batch")
    p_bat.add_argument("--base", required=True)
    p_bat.add_argument("--out", dest="outdir", required=True)
    p_bat.add_argument("--pattern", default="**/*.xlsx")
    p_bat.add_argument("--engine", choices=["auto","xlwings","headless"], default="auto")
    p_bat.add_argument("--to", choices=["csv","xlsx"], default="csv")
    p_bat.add_argument("--limit", type=int, default=None)

    p_saf = sub.add_parser("safety")
    p_saf.add_argument("--in", dest="infile", required=True)
    p_saf.add_argument("--abs", dest="abs_tol", type=float, default=1e-9)
    p_saf.add_argument("--rel", dest="rel_tol", type=float, default=1e-9)
    p_saf.add_argument("--limit", type=int, default=1000)
    p_saf.add_argument("--report", default=None)

    args = p.parse_args()
    if args.cmd == "export":
        if args.fmt == "csv":
            out = export_formulas_to_csv(args.infile, args.outfile, limit=args.limit, engine=args.engine)
        else:
            out = export_formulas_to_xlsx(args.infile, args.outfile, limit=args.limit, engine=args.engine)
        _print_json({"outfile": str(out)})

    elif args.cmd == "batch":
        res = batch_export(args.base, pattern=args.pattern, out_dir=args.outdir,
                           limit=args.limit, engine=args.engine, to=args.to)
        _print_json({"count": len(res), "outputs": [(str(i), str(o)) for i,o in res]})

    elif args.cmd == "safety":
        summ = safety_check_parity(args.infile, sample_limit=args.limit,
                                   abs_tol=args.abs_tol, rel_tol=args.rel_tol,
                                   out_report_csv=args.report)
        _print_json(summ)
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: class FormulaRow ---
@dataclass
class FormulaRow:
    sheet: str
    address: str   # z.B. "B7"
    formula: str   # inkl. '='
    value: Any     # berechneter Wert
    value_type: str  # human-readable Typ: number/text/date/bool/error/blank
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: class ExcelEngineXlwings ---
class ExcelEngineXlwings:
    """
    Nutzt Excel für Berechnung (maximale Übereinstimmung, erfordert Excel).
    """
    def __init__(self, path: str | Path):
        import xlwings as xw
        self._xw = xw
        self._app = xw.App(visible=False, add_book=False)
        try:
            self._wb = self._app.books.open(str(path), read_only=True, update_links=False)
        except Exception:
            self._wb = self._app.books.open(str(path), update_links=False)
        with contextlib.suppress(Exception):
            self._wb.api.Application.CalculateFullRebuild()
        try:
            self.date1904 = bool(self._wb.api.Date1904)
        except Exception:
            self.date1904 = False

    def value(self, sheet: str, address: str) -> Any:
        sht = self._wb.sheets[sheet]
        return sht.range(address).value

    def close(self):
        with contextlib.suppress(Exception):
            self._wb.close()
        with contextlib.suppress(Exception):
            self._app.quit()
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: class ExcelEngineHeadless ---
class ExcelEngineHeadless:
    """
    Headless-Auswertung mit xlcalculator (ohne Excel).
    """
    def __init__(self, path: str | Path):
        from xlcalculator import ModelCompiler, Model
        comp = ModelCompiler()
        parsed_model = comp.read_and_parse_archive(str(path))
        self._model = Model(parsed_model)

    def evaluate(self, ref: str) -> Any:
        return self._model.evaluate(ref)
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: class Mismatch ---
@dataclass
class Mismatch:
    ref: str
    excel_value: Any
    headless_value: Any
    abs_delta: Optional[float]
    rel_delta: Optional[float]
    note: str
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: method ExcelEngineXlwings.__init__ ---
    def __init__(self, path: str | Path):
        import xlwings as xw
        self._xw = xw
        self._app = xw.App(visible=False, add_book=False)
        try:
            self._wb = self._app.books.open(str(path), read_only=True, update_links=False)
        except Exception:
            self._wb = self._app.books.open(str(path), update_links=False)
        with contextlib.suppress(Exception):
            self._wb.api.Application.CalculateFullRebuild()
        try:
            self.date1904 = bool(self._wb.api.Date1904)
        except Exception:
            self.date1904 = False
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: method ExcelEngineXlwings.value ---
    def value(self, sheet: str, address: str) -> Any:
        sht = self._wb.sheets[sheet]
        return sht.range(address).value
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: method ExcelEngineXlwings.close ---
    def close(self):
        with contextlib.suppress(Exception):
            self._wb.close()
        with contextlib.suppress(Exception):
            self._app.quit()
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: method ExcelEngineHeadless.__init__ ---
    def __init__(self, path: str | Path):
        from xlcalculator import ModelCompiler, Model
        comp = ModelCompiler()
        parsed_model = comp.read_and_parse_archive(str(path))
        self._model = Model(parsed_model)
# --- DEF BLOCK END ---


# --- REQUIRED IMPORT SUGGESTIONS (dedupe on apply) ---
from __future__ import annotations
import math
import csv
import json
import contextlib
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple
from pathlib import Path

# --- DEF BLOCK START: method ExcelEngineHeadless.evaluate ---
    def evaluate(self, ref: str) -> Any:
        return self._model.evaluate(ref)
# --- DEF BLOCK END ---

